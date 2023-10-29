import xlrd
import xlwt
import pandas as pd
import openpyxl

guzi={}
cn=set()
ren={}

file = "test.xlsx"
xx=xlrd.open_workbook(file)
xs=xx.sheet_by_name("工作表2")

def FindTheHead():
    find=[]
    for row in range(0,xs.nrows):
        for col in range(0,xs.ncols):
            temp=xs.cell_value(row,col)

            if temp!="":
                find.append(row)
                find.append(col)

                return find

#创造字典存谷子
for i in range(FindTheHead()[1],xs.ncols):
    name = xs.cell_value(FindTheHead()[0], i)
    if xs.cell_value(FindTheHead()[0]+1, i)!='':
     price =round(float(xs.cell_value(FindTheHead()[0]+1, i)),2)
    guzi[name]=price
print(guzi)

#获得所有人
for i in range(FindTheHead()[1],xs.ncols):
    for j in range(FindTheHead()[0]+2,xs.nrows):
      temp = xs.cell_value(j,i)
      if temp!="":
       cn.add(temp)
print(cn)


#遍历

for i in range(FindTheHead()[1],xs.ncols):
    for j in range(FindTheHead()[0]+2,xs.nrows):
      name = xs.cell_value(j,i)
      someone = xs.cell_value(FindTheHead()[0], i)
      if name in cn :
          t=[]
          if someone not in t:
              if ren.get(name):
                t=ren.get(name)
                t.append(someone)
                ren[name]=t
              else:
                  t=[]
                  t.append(someone)
                  ren[name]=t
print(ren)

#计算每个cn在排表中的次数
count= {}
all=[]
for i in range(FindTheHead()[1],xs.ncols):
    for j in range(FindTheHead()[0]+2,xs.nrows):
        all.append(xs.cell_value(j,i))
for c in cn:
    count[c]=all.count(c)
print(count)
print(all.count(""))
print(len(all))
print(len(all)-all.count(""))
#合成pai的字符串
pai={}
for c in cn:
    a=""
    te = set(ren[c])
    for o in te:
       a=a+o+str(ren[c].count(o))
    pai[c]=a
print(pai)

#计算肾值
shen={}
for c in cn:
    money=0.0
    te = set(ren[c])
    for o in te:
       money=round(money+round(ren[c].count(o)*guzi[o],2),2)
    shen[c]=money
print(shen)

data1=pd.DataFrame.from_dict(pai,orient='index',columns=['排'])
data2=pd.DataFrame.from_dict(count,orient='index',columns=['总次数'])
data3=pd.DataFrame.from_dict(shen,orient='index',columns=['总肾'])
print(data1)
print(data2)
data=pd.concat([data1,data2],axis=1)
data=pd.concat([data,data3],axis=1)
print(data)


writer = pd.ExcelWriter(file,mode='a',engine='openpyxl',if_sheet_exists='new')

data.to_excel(writer,sheet_name='新表2')

writer.save()
writer.close()




class CXlAutofit():
    # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
    def get_num_colnum_dict(self):
        '''
        :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
        '''
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
        AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict

    # 自适应列宽
    def style_excel(self, excel_name: str, sheet_name: str):
        '''
        :param sheet_name:  excel中的sheet名
        :return:
        '''
        # 打开excel
        wb = openpyxl.load_workbook(excel_name)
        # 选择对应的sheet
        sheet = wb[sheet_name]
        # 获取最大行数与最大列数
        max_column = sheet.max_column
        max_row = sheet.max_row

        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
        num_str_dict = self.get_num_colnum_dict()

        # 遍历全部列
        for i in range(1, max_column + 1):
            # 遍历每一列的全部行
            for j in range(1, max_row + 1):
                column = 0
                # 获取j行i列的值
                sheet_value = sheet.cell(row=j, column=i).value
                # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                sheet_value_list = [k for k in str(sheet_value)]
                # 遍历当前单元格的字符列表
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    if v.isdigit() == True or v.isalpha() == True:
                        column += 2
                    else:
                        column += 4
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                try:
                    if column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[num_str_dict[key]].width = value
        # 保存
        wb.save(excel_name)


# 调用方法 实例化类
Entity = CXlAutofit()
# 传入参数：Excel名称，需要设置列宽的Sheet名称
Entity.style_excel(file, '新表2')
